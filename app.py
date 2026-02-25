from flask import Flask, render_template, request, redirect, url_for, session, flash
import mysql.connector
from werkzeug.security import generate_password_hash, check_password_hash
import io
from flask import send_file, flash, redirect, url_for
from openpyxl import Workbook
import uuid, time
from flask import make_response
from flask import session
from datetime import datetime, timedelta
from flask import jsonify
import hashlib
from datetime import datetime, timedelta
import secrets
from datetime import date
import os
from flask import send_from_directory



# Dictionary to store active QR sessions
active_qr_sessions = {}

def generate_session_token(user_id):
    """Generate a unique session token"""
    return f"{user_id}_{secrets.token_hex(16)}"

def create_qr_session(user_id, subject, max_scans, time_limit):
    """Create a new QR session"""
    session_id = generate_session_token(user_id)
    expires_at = datetime.now() + timedelta(seconds=time_limit)
    
    active_qr_sessions[session_id] = {
        'user_id': user_id,
        'subject': subject,
        'max_scans': max_scans,
        'expires_at': expires_at,
        'created_at': datetime.now(),
        'devices': set()
    }
    
    return session_id

def validate_qr_session(session_id, fingerprint):
    """Check if a QR session is valid"""
    if session_id not in active_qr_sessions:
        return False, "Invalid QR code"
    
    session = active_qr_sessions[session_id]
    
    # Check expiration
    if datetime.now() > session['expires_at']:
        del active_qr_sessions[session_id]
        return False, "QR code expired"
    
    # Check scan limit
    if session['scans_used'] >= session['max_scans']:
        return False, "Scan limit reached"
    
    # Check if device already scanned
    if fingerprint in session['devices']:
        return False, "Device already scanned"
    
    return True, ""
# ---------------------
qr_scan_tracker = {} 
app = Flask(__name__)
app.secret_key = 'your_secret_key'

# MySQL Configuration
def get_db_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="Kito@0512",
        database="user_portal"
    )


def get_client_fingerprint():
    """Generate a fingerprint based on user agent and IP"""
    user_agent = request.headers.get('User-Agent', '')
    ip = request.remote_addr
    fingerprint_string = f"{user_agent}{ip}"
    return hashlib.md5(fingerprint_string.encode()).hexdigest()

def nocache(view):
    def no_cache(*args, **kwargs):
        resp = make_response(view(*args, **kwargs))
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
        return resp
    no_cache.__name__ = view.__name__
    return no_cache

@app.route('/')
def home():
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        hashed_password = generate_password_hash(password)

        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        cursor.execute("SELECT * FROM users WHERE email=%s", (email,))
        if cursor.fetchone():
            flash("Email already registered")
            return redirect(url_for('register'))

        cursor.execute("INSERT INTO users (username, email, password_hash) VALUES (%s, %s, %s)",
                       (username, email, hashed_password))
        db.commit()
        cursor.close()
        db.close()
        flash("Registered successfully! Please login.")
        return redirect(url_for('home'))
    return render_template('register.html')

@app.route('/login', methods=['POST'])
def login():
    email = request.form['email']
    password = request.form['password']

    db = get_db_connection()
    cursor = db.cursor(dictionary=True)

    cursor.execute("SELECT * FROM users WHERE email=%s", (email,))
    user = cursor.fetchone()
    if user and check_password_hash(user['password_hash'], password):
        session['user'] = user['username']
        session['user_id'] = user['id']  # Assuming your users table has an 'id' column
        return redirect(url_for('dashboard'))
    else:
        flash("Invalid email or password")
        return redirect(url_for('home'))

@app.route('/dashboard')
def dashboard():
    if 'user' in session:
        return render_template('dashboard.html', 
                             username=session['user'],
                             user_id=session.get('user_id'))  # user_id भी pass करें
    else:
        flash("Please log in first")
        return redirect(url_for('home'))

@app.route('/logout')
def logout():
    session.pop('user', None)
    session.pop('user_id', None)
    flash("Logged out successfully")
    return redirect(url_for('home'))

@app.route('/get-profile')
def get_profile():
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
        
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        cursor.execute("SELECT username, email FROM users WHERE id=%s", (session['user_id'],))
        user = cursor.fetchone()
        
        cursor.close()
        db.close()

        return jsonify({
            'success': True,
            'username': user['username'],
            'email': user['email']
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/update-profile', methods=['POST'])
def update_profile():
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
        
    data = request.get_json()
    new_username = data.get('username')
    email = data.get('email')
    role = data.get('role')
    subject = data.get('subject')
    
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        # Check if username already exists (excluding current user)
        cursor.execute("SELECT * FROM users WHERE username=%s AND id != %s", (new_username, session['user_id']))
        if cursor.fetchone():
            return jsonify({'error': 'Username already exists'}), 400

        # Update user profile
        cursor.execute(
            "UPDATE users SET username=%s, email=%s WHERE id=%s",
            (new_username, email, session['user_id'])
        )
        db.commit()
        cursor.close()
        db.close()

        # Update session
        session['user'] = new_username
        
        return jsonify({
            'success': True, 
            'message': 'Profile updated successfully',
            'new_username': new_username
        })
        
    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/delete_user/<username>', methods=['POST'])
def delete_user(username):
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)

    cursor.execute("DELETE FROM users WHERE username=%s", (username,))
    db.commit()
    cursor.close()
    db.close()

    if session.get('username') == username:
        session.pop('username', None)
    
    flash('User deleted successfully')
    return redirect(url_for('home'))

# Dictionary to store scan counts and device fingerprints
scan_counts = {}
device_fingerprints = {}

@app.route('/attendance-form')
@nocache
def attendance_form():

    subject = request.args.get('subject')
    max_scans = int(request.args.get('maxScans'))
    time_limit = int(request.args.get('timeLimit'))
    
    # Create a unique key for this QR session
    qr_key = f"{subject}_{session.get('user_id')}"
    
    # Get client fingerprint
    fingerprint = get_client_fingerprint()
    
    # Clean up old entries (older than 24 hours)
    now = datetime.now()
    for key in list(scan_counts.keys()):
        if 'timestamp' in scan_counts[key] and (now - scan_counts[key]['timestamp']) > timedelta(hours=24):
            del scan_counts[key]
            if key in device_fingerprints:
                del device_fingerprints[key]
    
    # Initialize scan count if not exists
    if qr_key not in scan_counts:
        scan_counts[qr_key] = {
            'count': 0,
            'timestamp': now,
            'max_scans': max_scans
        }
        device_fingerprints[qr_key] = set()
    # If exists but max_scans is different, reset count
    elif scan_counts[qr_key]['max_scans'] != max_scans:
        scan_counts[qr_key] = {
            'count': 0,
            'timestamp': now,
            'max_scans': max_scans
        }
        device_fingerprints[qr_key] = set()
    
    # Check if scan limit reached
    if scan_counts[qr_key]['count'] >= max_scans:
        return render_template('scan_limit_reached.html')
    
    # Check if this device already scanned
    if fingerprint in device_fingerprints[qr_key]:
        return render_template('device_already_scanned.html')
    
    return render_template('attendance_form.html', 
                         subject=subject,
                         max_scans=max_scans,
                         time_limit=time_limit)

@app.route('/msg')
def msg():
    return render_template('msg.html')

@app.route('/submit-attendance', methods=['POST'])
def submit_attendance():
    subject = request.form.get('subject')
    name = request.form.get('name')
    rollno = request.form.get('rollno')
    
    qr_key = f"{subject}_{session.get('user_id')}"
    fingerprint = get_client_fingerprint()
    
    # Check if this device already scanned (in case they bypassed the form check)
    if qr_key in device_fingerprints and fingerprint in device_fingerprints[qr_key]:
        flash("❌ You have already submitted attendance from this device", "error")
        return redirect(url_for('msg'))
    
    # Increment scan count and add fingerprint
    if qr_key in scan_counts:
        scan_counts[qr_key]['count'] += 1
        device_fingerprints[qr_key].add(fingerprint)
        sessions = qr_sessions.get(session.get('user_id'), [])
        for s in sessions:
            if s['subject'] == subject:
                s['scans_used'] = scan_counts[qr_key]['count']
                if s['scans_used'] >= s['max_scans']:
                    s['status'] = 'Expired'
    
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        cursor.execute("INSERT INTO attendance (name, rollno, subject) VALUES (%s, %s, %s)", 
                       (name, rollno, subject))
        db.commit()
        cursor.close()
        db.close()
        flash("✅ Attendance submitted successfully!", "success")
    except Exception as e:
        db.rollback()
        flash(f"❌ Error saving attendance: {e}", "error")

    return redirect(url_for('msg'))

import uuid

@app.route('/generate-qr', methods=['POST'])
def generate_qr():
    subject = request.form['subject']
    max_scans = int(request.form['maxScans'])

    qr_id = str(uuid.uuid4())  # unique ID for QR

    qr_scan_data[qr_id] = {'max': max_scans, 'count': 0}

    qr_url = f"http://localhost:5000/attendance-form?qr_id={qr_id}"
    # Generate QR with qr_url and return

    return render_template('qr_generated.html', qr_url=qr_url)

# Key = QR ID, Value = {'max': int, 'count': int}






# Add this route to your Flask app
@app.route('/save-session', methods=['POST'])
def save_session():
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
        
    data = request.get_json()
    subject = data.get('subject')
    time_limit = data.get('timeLimit')
    max_scans = data.get('maxScans')
    class_name = data.get('class_name')
    
    # Generate a unique session ID
    session_id = f"{session['user_id']}_{secrets.token_hex(16)}"
    
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)
        # Check if session_id column exists
        cursor.execute("SHOW COLUMNS FROM qr_sessions LIKE 'session_id'")
        session_id_exists = cursor.fetchone()
        
        if session_id_exists:
            # Insert with session_id
            cursor.execute(
                "INSERT INTO qr_sessions (user_id, session_id, subject, time_limit, max_scans, class_name, scans_used) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                (session['user_id'], session_id, subject, time_limit, max_scans, class_name, 0)
            )
        else:
            # Insert without session_id (use auto-increment ID)
            cursor.execute(
                "INSERT INTO qr_sessions (user_id, subject, time_limit, max_scans, class_name, scans_used) VALUES (%s, %s, %s, %s, %s, %s)",
                (session['user_id'], subject, time_limit, max_scans, class_name, 0)
            )
            # Get the auto-generated ID
            cursor.execute("SELECT LAST_INSERT_ID() as id")
            result = cursor.fetchone()
            session_id = str(result['id'])
        
        db.commit()
        cursor.close()
        db.close()

        
        # Store in active sessions
        expires_at = datetime.now() + timedelta(seconds=int(time_limit))
        active_qr_sessions[session_id] = {
            'user_id': session['user_id'],
            'subject': subject,
            'max_scans': max_scans,
            'expires_at': expires_at,
            'created_at': datetime.now(),
            'scans_used': 0,
            'devices': set(),
            'class_name': class_name
        }
        
        return jsonify({
            'success': True,
            'session_id': session_id,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500

# Add this route to fetch sessions
@app.route('/delete-session/<session_id>', methods=['DELETE'])
def delete_session(session_id):
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        # Delete the session from database
        cursor.execute("DELETE FROM qr_sessions WHERE session_id = %s AND user_id = %s", 
                      (session_id, session['user_id']))
        
        # Also delete any attendance records associated with this session
        cursor.execute("DELETE FROM attendance WHERE session_id = %s", (session_id,))
        
        db.commit()
        cursor.close()
        db.close()

        
        # Remove from active sessions if exists
        if session_id in active_qr_sessions:
            del active_qr_sessions[session_id]
        
        return jsonify({'success': True, 'message': 'Session deleted successfully'})
    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500


#students-----------------------------
# Add these routes to app.py

@app.route('/add-student', methods=['POST'])
def add_student():
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
        
    data = request.get_json()
    rollno = data.get('rollno')
    name = data.get('name')
    student_class = data.get('class')
    stream = data.get('stream')
    
    # Determine which table to use based on class
    if student_class.startswith('FY'):
        table_name = 'students_fy'
    elif student_class.startswith('SY'):
        table_name = 'students_sy'
    elif student_class.startswith('TY'):
        table_name = 'students_ty'
    else:
        return jsonify({'error': 'Invalid class'}), 400
    
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        cursor.execute(
            f"INSERT INTO {table_name} (rollno, name, class, stream) VALUES (%s, %s, %s, %s)",
            (rollno, name, student_class, stream)
        )
        db.commit()
        cursor.close()
        db.close()
        return jsonify({'success': True, 'message': 'Student added successfully'})
    except mysql.connector.IntegrityError:
        db.rollback()
        return jsonify({'error': 'Student with this roll number already exists'}), 400
    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/get-students/<class_type>')
def get_students(class_type):
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
        
    # Determine which table to use based on class type
    if class_type == 'fy':
        table_name = 'students_fy'
    elif class_type == 'sy':
        table_name = 'students_sy'
    elif class_type == 'ty':
        table_name = 'students_ty'
    else:
        return jsonify({'error': 'Invalid class type'}), 400
    
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        cursor.execute(f"SELECT * FROM {table_name} ORDER BY rollno")
        students = cursor.fetchall()
        
        # Convert datetime objects to strings for JSON serialization
        for student in students:
            if isinstance(student['created_at'], datetime):
                student['created_at'] = student['created_at'].strftime('%Y-%m-%d %H:%M:%S')
                
        return jsonify({'students': students})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/delete-student/<class_type>/<rollno>', methods=['DELETE'])
def delete_student(class_type, rollno):
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
        
    # Determine which table to use based on class type
    if class_type == 'fy':
        table_name = 'students_fy'
    elif class_type == 'sy':
        table_name = 'students_sy'
    elif class_type == 'ty':
        table_name = 'students_ty'
    else:
        return jsonify({'error': 'Invalid class type'}), 400
    
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        cursor.execute(f"DELETE FROM {table_name} WHERE rollno = %s", (rollno,))
        db.commit()
        cursor.close()
        db.close()
        
        return jsonify({'success': True, 'message': 'Student deleted successfully'})
    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/download-student-list/<class_type>')
def download_student_list(class_type):
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
        
    stream_filter = request.args.get('stream', 'all')
    
    # Determine which tables to use based on class type
    tables_to_query = []
    if class_type == 'all':
        tables_to_query = ['students_fy', 'students_sy', 'students_ty']
        class_display = 'All Classes'
    elif class_type == 'fy':
        tables_to_query = ['students_fy']
        class_display = 'FY'
    elif class_type == 'sy':
        tables_to_query = ['students_sy']
        class_display = 'SY'
    elif class_type == 'ty':
        tables_to_query = ['students_ty']
        class_display = 'TY'
    else:
        return jsonify({'error': 'Invalid class type'}), 400
    
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        all_students = []
        
        for table_name in tables_to_query:
            if stream_filter != 'all':
                cursor.execute(f"SELECT rollno, name, class, stream FROM {table_name} WHERE LOWER(stream) = LOWER(%s) ORDER BY rollno", (stream_filter,))
            else:
                cursor.execute(f"SELECT rollno, name, class, stream FROM {table_name} ORDER BY rollno")
            
            students = cursor.fetchall()
            all_students.extend(students)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        
        # Set title based on filters
        if stream_filter != 'all':
            ws.title = f"{class_display} {stream_filter.upper()} Students"
            title = f"Class: {class_display} | Stream: {stream_filter.upper()}"
        else:
            ws.title = f"{class_display} Student List"
            title = f"Class: {class_display}"
        
        # Add headers
        ws['A1'] = title
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Add column headers
        ws['A4'] = "Roll No"
        ws['B4'] = "Name"
        ws['C4'] = "Class"
        ws['D4'] = "Stream"
        
        # Add student data
        for i, student in enumerate(all_students, start=5):
            ws[f'A{i}'] = student['rollno']
            ws[f'B{i}'] = student['name']
            ws[f'C{i}'] = student['class']
            ws[f'D{i}'] = student['stream']
        
        # Create a BytesIO buffer and save the workbook to it
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        cursor.close()
        db.close()
        
        # Return the Excel file
        if stream_filter != 'all':
            filename = f"student_list_{class_display}_{stream_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            filename = f"student_list_{class_display}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500
        
# Update the get-sessions route to include status calculation
@app.route('/get-sessions')
def get_sessions():
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
        
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        cursor.execute(
            "SELECT session_id, subject, time_limit, max_scans, created_at, class_name FROM qr_sessions WHERE user_id = %s ORDER BY created_at DESC",
            (session['user_id'],)
        )
        sessions = cursor.fetchall()
        
        # Add status to each session
        for session_data in sessions:
            if isinstance(session_data['created_at'], datetime):
                session_data['created_at'] = session_data['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            
            # Calculate status based on time limit and creation time
            created_time = datetime.strptime(session_data['created_at'], '%Y-%m-%d %H:%M:%S')
            current_time = datetime.now()
            time_elapsed = (current_time - created_time).total_seconds()
            
            if session_data['time_limit'] and time_elapsed > session_data['time_limit']:
                session_data['status'] = 'expired'
            else:
                session_data['status'] = 'active'
                
        return jsonify({'sessions': sessions})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
qr_scan_data = {}
qr_sessions = {}

@app.route('/get-session-dates')
def get_session_dates():
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
        
    subject = request.args.get('subject')
    class_name = request.args.get('class_name')
    
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        cursor.execute(
            "SELECT DISTINCT DATE(created_at) as session_date FROM qr_sessions WHERE user_id = %s AND subject = %s AND class_name = %s ORDER BY session_date DESC",
            (session['user_id'], subject, class_name)
        )
        dates = cursor.fetchall()
        
        date_list = [date['session_date'].strftime('%Y-%m-%d') for date in dates if date['session_date']]
        
        return jsonify({'dates': date_list})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/download-default-absent-list/<class_name>')
def download_default_absent_list(class_name):
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    
    # Get stream and session parameters
    stream = request.args.get('stream', '')
    subject = request.args.get('subject', '')
    session_date = request.args.get('session_date', '')
    
    if not subject or not session_date:
        return jsonify({'error': 'Subject and session date are required'}), 400
    
    # Determine which table to use based on class
    if class_name.startswith('FY'):
        table_name = 'students_fy'
    elif class_name.startswith('SY'):
        table_name = 'students_sy'
    elif class_name.startswith('TY'):
        table_name = 'students_ty'
    else:
        return jsonify({'error': 'Invalid class'}), 400
    
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        # Get ALL students for this class and stream
        if stream:
            cursor.execute(f"SELECT rollno, name, stream FROM {table_name} WHERE stream = %s ORDER BY rollno", (stream,))
        else:
            cursor.execute(f"SELECT rollno, name, stream FROM {table_name} ORDER BY rollno")
        
        all_students = cursor.fetchall()
        
        # Get PRESENT students for this subject and session date
        cursor.execute(
            "SELECT rollno FROM attendance WHERE subject = %s AND DATE(created_at) = %s",
            (subject, session_date)
        )
        present_students = {student['rollno'] for student in cursor.fetchall()}
        
        # Filter ABSENT students only (those not in present_students)
        absent_students = [
            student for student in all_students 
            if student['rollno'] not in present_students
        ]
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Absent Students List"
        
        # Add headers
        ws['A1'] = f"Subject: {subject}"
        ws['A2'] = f"Class: {class_name}"
        if stream:
            ws['A3'] = f"Stream: {stream}"
        ws['A4'] = f"Session Date: {session_date}"
        ws['A5'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A6'] = f"Total Absent Students: {len(absent_students)}"
        
        # Add column headers
        ws['A8'] = "Roll No"
        ws['B8'] = "Name"
        ws['C8'] = "Stream"
        ws['D8'] = "Status"
        
        # Add absent student data
        for i, student in enumerate(absent_students, start=9):
            ws[f'A{i}'] = student['rollno']
            ws[f'B{i}'] = student['name']
            ws[f'C{i}'] = student.get('stream', '')
            ws[f'D{i}'] = "Absent"
        
        # Create a BytesIO buffer and save the workbook to it
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        cursor.close()
        db.close()
        
        # Return the Excel file
        if stream:
            filename = f"absent_list_{subject}_{class_name}_{stream}_{session_date.replace('-', '')}.xlsx"
        else:
            filename = f"absent_list_{subject}_{class_name}_{session_date.replace('-', '')}.xlsx"
            
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Update the generate-report function to use session date but store in filename
@app.route('/generate-report', methods=['POST'])
def generate_report():
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
        
    data = request.get_json()
    subject = data.get('subject')
    class_name = data.get('class_name')
    session_date = data.get('session_date')
    
    # Use session date instead of current date for filtering
    target_date = session_date if session_date else date.today().strftime('%Y-%m-%d')
    
    # Extract stream from class_name (e.g., FYBBA -> BBA, FYBBACA -> BBACA)
    stream = class_name[2:]  # Remove FY/SY/TY prefix
    
    # Determine which table to use based on class
    if class_name.startswith('FY'):
        table_name = 'students_fy'
    elif class_name.startswith('SY'):
        table_name = 'students_sy'
    elif class_name.startswith('TY'):
        table_name = 'students_ty'
    else:
        return jsonify({'error': 'Invalid class'}), 400
    
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        # Get students for this class and specific stream
        cursor.execute(f"SELECT rollno, name FROM {table_name} WHERE stream = %s ORDER BY rollno", (stream,))
        students = cursor.fetchall()
        
        # Get attendance data for this subject and class on the specific session date
        cursor.execute(
            "SELECT rollno FROM attendance WHERE subject = %s AND DATE(created_at) = %s",
            (subject, target_date)
        )
        present_students = {student['rollno'] for student in cursor.fetchall()}
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Add headers with session date
        ws['A1'] = f"Subject: {subject}"
        ws['A2'] = f"Class: {class_name}"
        ws['A3'] = f"Stream: {stream}"
        ws['A4'] = f"Session Date: {target_date}"
        ws['A5'] = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Add column headers
        ws['A7'] = "Roll No"
        ws['B7'] = "Name"
        ws['C7'] = "Stream"
        ws['D7'] = "Status"
        
        # Add student data
        for i, student in enumerate(students, start=8):
            ws[f'A{i}'] = student['rollno']
            ws[f'B{i}'] = student['name']
            ws[f'C{i}'] = stream
            ws[f'D{i}'] = "Present" if student['rollno'] in present_students else "Absent"
        
        # Save the workbook - include session date in filename
        filename = f"report_{subject}_{class_name}_{target_date.replace('-', '')}_{secrets.token_hex(8)}.xlsx"
        filepath = os.path.join('reports', filename)
        wb.save(filepath)
        
        # Save report info to database
        cursor.execute(
            "INSERT INTO reports (user_id, filename, subject, class_name, report_date) VALUES (%s, %s, %s, %s, %s)",
            (session['user_id'], filename, subject, class_name, date.today())
        )
        db.commit()
        cursor.close()
        db.close()
        
        return jsonify({'success': True, 'filename': filename, 'session_date': target_date})
    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/get-reports')
def get_reports():
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
        
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        cursor.execute(
            "SELECT * FROM reports WHERE user_id = %s ORDER BY created_at DESC",
            (session['user_id'],)
        )
        reports = cursor.fetchall()
        
        # Convert datetime objects to strings for JSON serialization
        for report in reports:
            if isinstance(report['created_at'], datetime):
                report['created_at'] = report['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            if isinstance(report['report_date'], date):
                report['report_date'] = report['report_date'].strftime('%Y-%m-%d')
                
        return jsonify({'reports': reports})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/download-report/<filename>')
def download_report(filename):
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    
    # Verify the user owns this report
    cursor.execute(
        "SELECT * FROM reports WHERE filename = %s AND user_id = %s",
        (filename, session['user_id'])
    )
    report = cursor.fetchone()
    
    if not report:
        return jsonify({'error': 'Report not found'}), 404
        
    return send_from_directory('reports', filename, as_attachment=True)


@app.route('/delete-report/<filename>', methods=['DELETE'])
def delete_report(filename):
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
        
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        # Verify the user owns this report
        cursor.execute(
            "SELECT * FROM reports WHERE filename = %s AND user_id = %s",
            (filename, session['user_id'])
        )
        report = cursor.fetchone()
        
        if not report:
            cursor.close()
            db.close()
            return jsonify({'error': 'Report not found'}), 404
        
        # Delete from database
        cursor.execute(
            "DELETE FROM reports WHERE filename = %s AND user_id = %s",
            (filename, session['user_id'])
        )
        db.commit()
        cursor.close()
        db.close()
        
        # Delete the file
        filepath = os.path.join('reports', filename)
        if os.path.exists(filepath):
            os.remove(filepath)
            
        return jsonify({'success': True})
    except Exception as e:
        db.rollback()
        cursor.close()
        db.close()
        return jsonify({'error': str(e)}), 500

# Subjects management routes
@app.route('/add-subject', methods=['POST'])
def add_subject():
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
        
    data = request.get_json()
    class_name = data.get('class_name')
    subject_name = data.get('subject_name')
    
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        cursor.execute(
            "INSERT INTO subjects (user_id, class_name, subject_name) VALUES (%s, %s, %s)",
            (session['user_id'], class_name, subject_name)
        )
        db.commit()
        cursor.close()
        db.close()
        return jsonify({'success': True, 'message': 'Subject added successfully'})
    except mysql.connector.IntegrityError:
        db.rollback()
        return jsonify({'error': 'Subject already exists for this class'}), 400
    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/get-subjects/<class_name>')
def get_subjects(class_name):
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
        
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        cursor.execute(
            "SELECT * FROM subjects WHERE user_id = %s AND class_name = %s ORDER BY subject_name",
            (session['user_id'], class_name)
        )
        subjects = cursor.fetchall()
        
        # Convert datetime objects to strings for JSON serialization
        for subject in subjects:
            if isinstance(subject['created_at'], datetime):
                subject['created_at'] = subject['created_at'].strftime('%Y-%m-%d %H:%M:%S')
                
        return jsonify({'subjects': subjects})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/delete-subject/<int:subject_id>', methods=['DELETE'])
def delete_subject(subject_id):
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
        
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        cursor.execute(
            "DELETE FROM subjects WHERE id = %s AND user_id = %s",
            (subject_id, session['user_id'])
        )
        db.commit()
        cursor.close()
        db.close()
        return jsonify({'success': True, 'message': 'Subject deleted successfully'})
    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/get-classes')
def get_classes():
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
        
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        cursor.execute(
            "SELECT DISTINCT class_name FROM subjects WHERE user_id = %s ORDER BY class_name",
            (session['user_id'],)
        )
        classes = cursor.fetchall()
        class_names = [cls['class_name'] for cls in classes]
        return jsonify({'classes': class_names})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
        
if __name__ == '__main__':
    app.run(host="0.0.0.0", port=5000, debug=True)

